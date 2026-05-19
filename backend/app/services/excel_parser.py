"""Parse an .xlsx register-definition file.

Sheet name: Registers
Header row: 7 (1-indexed, i.e. row index 6 in openpyxl)
Columns: ADDR | Register | INI | Bits | Member
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Tuple, Union

import openpyxl


@dataclass
class BitFieldInfo:
    name: str
    low_bit: int
    width: int
    ini: str  # raw hex string from Excel, e.g. "0x0"
    register_name: str
    register_addr: str  # 4-char uppercase hex, e.g. "0010"


@dataclass
class RegisterInfo:
    addr: str  # 4-char uppercase hex
    name: str
    bitfields: list[BitFieldInfo] = field(default_factory=list)


def _parse_bits(bits_str: str) -> tuple[int, int]:
    """Return (low_bit, width) from a Bits cell value like '5_4' or '10'."""
    s = str(bits_str).strip()
    if "_" in s:
        parts = s.split("_")
        hi = int(parts[0])
        lo = int(parts[1])
        return lo, hi - lo + 1
    else:
        bit = int(s)
        return bit, 1


def parse_excel(file_path: Union[str, Path]) -> Tuple[Dict[str, RegisterInfo], List[BitFieldInfo]]:
    """Parse Excel and return (registers_by_addr, ordered_bitfields).

    registers_by_addr maps uppercase 4-char addr -> RegisterInfo.
    ordered_bitfields is a flat list in address/definition order (used as column order).
    """
    # read_only=False: avoids XML-stream-consumed issue when BytesIO is iterated twice
    wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)

    # Search all sheets for one containing a proper header row:
    # must have "ADDR" AND at least one companion column in the same row.
    # This avoids false-positives where "ADDR" appears as a data value or title.
    _COMPANION_COLS = {"REGISTER", "BITS", "MEMBER", "INI"}
    target_sheet_name = None
    header_from_scan: int | None = None

    for ws_candidate in wb.worksheets:
        peek = list(ws_candidate.iter_rows(values_only=True, max_row=50))
        print(f"[excel_parser] scanning sheet={ws_candidate.title!r}, peek_rows={len(peek)}")
        for row_idx, row in enumerate(peek):
            if not row:
                continue
            cells_upper = {
                str(c).strip().upper()
                for c in row
                if c is not None and str(c).strip()
            }
            if cells_upper:
                print(f"[excel_parser]   row{row_idx} non-empty cells: {cells_upper}")
            if "ADDR" in cells_upper and cells_upper & _COMPANION_COLS:
                target_sheet_name = ws_candidate.title
                header_from_scan = row_idx
                print(f"[excel_parser]   => MATCH: header at row {row_idx}")
                break
        if target_sheet_name:
            break

    if target_sheet_name is None:
        print(f"[excel_parser] WARNING: no sheet found with ADDR + companion column; falling back to active sheet")

    ws = wb[target_sheet_name] if target_sheet_name else wb.active

    print(f"[excel_parser] sheets={[s.title for s in wb.worksheets]}, selected={ws.title!r}")

    rows = list(ws.iter_rows(values_only=True))
    print(f"[excel_parser] total rows={len(rows)}")

    # Use the header row found during the sheet scan when available;
    # otherwise fall back to scanning rows (handles headers beyond row 50).
    if header_from_scan is not None:
        header_row_idx = header_from_scan
    else:
        header_row_idx = None
        for i, row in enumerate(rows):
            if not row:
                continue
            cells_upper = {
                str(c).strip().upper()
                for c in row
                if c is not None and str(c).strip()
            }
            if "ADDR" in cells_upper and cells_upper & _COMPANION_COLS:
                header_row_idx = i
                break
        if header_row_idx is None:
            header_row_idx = min(6, len(rows) - 1)

    if header_row_idx < 0 or header_row_idx >= len(rows):
        raise ValueError(f"Cannot find header row with ADDR column (sheet has {len(rows)} rows)")

    col_names = [str(c).strip().upper() if c is not None else "" for c in rows[header_row_idx]]

    def col(name: str) -> int:
        for i, c in enumerate(col_names):
            if c == name:
                return i
        return -1

    addr_col = col("ADDR")
    reg_col = col("REGISTER")
    ini_col = col("INI")
    bits_col = col("BITS")
    member_col = col("MEMBER")

    print(f"[excel_parser] header_row={header_row_idx}, cols: ADDR={addr_col} REGISTER={reg_col} INI={ini_col} BITS={bits_col} MEMBER={member_col}")
    if header_row_idx is not None and header_row_idx < len(rows):
        print(f"[excel_parser] header_row_content={rows[header_row_idx]}")

    registers: dict[str, RegisterInfo] = {}
    ordered_bitfields: list[BitFieldInfo] = []
    current_addr: str | None = None

    _EMPTY_STRINGS = {"none", "null", "n/a", "na", ""}

    def _cell(row: tuple, idx: int):
        return row[idx] if idx >= 0 and idx < len(row) else None

    def _val(raw) -> str | None:
        """Return None if the cell is empty or a known placeholder string."""
        if raw is None:
            return None
        s = str(raw).strip()
        if s.lower() in _EMPTY_STRINGS:
            return None
        return s

    for row in rows[header_row_idx + 1:]:
        # Skip completely empty rows
        if not row or all(c is None for c in row):
            continue

        addr_val = _val(_cell(row, addr_col))
        reg_val  = _val(_cell(row, reg_col))
        ini_raw  = _cell(row, ini_col)
        bits_raw = _cell(row, bits_col)
        member_val = _val(_cell(row, member_col))

        # Treat row as empty if all meaningful columns are placeholder
        if addr_val is None and reg_val is None and member_val is None:
            continue

        ini_val  = None if ini_raw  is None else str(ini_raw).strip()
        bits_val = None if bits_raw is None else str(bits_raw).strip()

        # New register block
        if addr_val:
            raw_addr = addr_val.upper()
            # Normalise to 4 hex chars, stripping any leading 0x
            raw_addr = raw_addr.lstrip("0X").lstrip("0X")  # handle "0x" prefix
            if raw_addr == "":
                raw_addr = "0"
            current_addr = raw_addr.zfill(4)

        if reg_val and current_addr is not None:
            if current_addr not in registers:
                registers[current_addr] = RegisterInfo(addr=current_addr, name=reg_val)

        if member_val and bits_val and current_addr is not None:
            try:
                low_bit, width = _parse_bits(str(bits_val))
            except (ValueError, IndexError):
                continue  # skip malformed rows

            bf = BitFieldInfo(
                name=member_val,
                low_bit=low_bit,
                width=width,
                ini=ini_val if ini_val is not None else "0x0",
                register_name=registers[current_addr].name if current_addr in registers else "",
                register_addr=current_addr,
            )
            if current_addr in registers:
                registers[current_addr].bitfields.append(bf)
            ordered_bitfields.append(bf)

    wb.close()
    print(f"[excel_parser] result: {len(registers)} registers, {len(ordered_bitfields)} bitfields")
    return registers, ordered_bitfields
