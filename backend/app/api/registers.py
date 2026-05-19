import io
from pathlib import Path
from datetime import datetime

import xlrd
import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from sqlalchemy.orm import Session

from ..db import get_db, DATA_DIR
from ..models import RegisterDefinitionORM, RegisterDefinitionOut, RegisterRenameIn
from ..services.excel_parser import parse_excel

router = APIRouter(prefix="/api/registers", tags=["registers"])


def _xls_to_xlsx(xls_bytes: bytes) -> bytes:
    """Convert legacy .xls bytes to .xlsx bytes, preserving cell values."""
    wb_xls = xlrd.open_workbook(file_contents=xls_bytes)
    print(f"[xls_to_xlsx] nsheets={wb_xls.nsheets}, names={wb_xls.sheet_names()}")
    wb_xlsx = openpyxl.Workbook()
    for sheet_idx in range(wb_xls.nsheets):
        ws_xls = wb_xls.sheet_by_index(sheet_idx)
        vis_list = getattr(wb_xls, 'sheet_visibility', None)
        visibility = vis_list[sheet_idx] if vis_list else 0  # 0=visible,1=hidden,2=very hidden
        print(f"[xls_to_xlsx] sheet[{sheet_idx}] name={ws_xls.name!r} visibility={visibility} nrows={ws_xls.nrows} ncols={ws_xls.ncols}")
        ws_xlsx = wb_xlsx.active if sheet_idx == 0 else wb_xlsx.create_sheet()
        ws_xlsx.title = ws_xls.name
        for row in range(ws_xls.nrows):
            for col_idx in range(ws_xls.ncols):
                cell = ws_xls.cell(row, col_idx)
                if cell.ctype == xlrd.XL_CELL_NUMBER:
                    val = cell.value
                    val = int(val) if val == int(val) else val
                elif cell.ctype == xlrd.XL_CELL_EMPTY:
                    val = None
                else:
                    val = cell.value
                ws_xlsx.cell(row=row + 1, column=col_idx + 1, value=val)
        # Print first non-empty rows for diagnosis
        sample_rows = []
        for r in range(min(10, ws_xls.nrows)):
            row_vals = [ws_xls.cell(r, c).value for c in range(ws_xls.ncols)]
            if any(v for v in row_vals if v is not None and str(v).strip()):
                sample_rows.append(f"    row{r}: {row_vals}")
            if len(sample_rows) >= 3:
                break
        if sample_rows:
            print(f"[xls_to_xlsx] first non-empty rows of sheet {ws_xls.name!r}:")
            for s in sample_rows:
                print(s)
    buf = io.BytesIO()
    wb_xlsx.save(buf)
    print(f"[xls_to_xlsx] conversion done, xlsx size={len(buf.getvalue())} bytes")
    return buf.getvalue()


@router.get("", response_model=list[RegisterDefinitionOut])
def list_registers(db: Session = Depends(get_db)):
    records = db.query(RegisterDefinitionORM).order_by(RegisterDefinitionORM.uploaded_at.desc()).all()
    return records


@router.post("", response_model=RegisterDefinitionOut)
async def upload_register(
    file: UploadFile = File(...),
    name: str = Form(""),
    db: Session = Depends(get_db),
):
    fname = file.filename or ""
    if not fname.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are accepted")

    content = await file.read()

    # Detect actual format by magic bytes (ignore extension — some xlsx are OLE2 internally)
    _OLE2_MAGIC = b'\xD0\xCF\x11\xE0'
    _ZIP_MAGIC  = b'PK\x03\x04'
    actual_fmt = 'xls' if content[:4] == _OLE2_MAGIC else ('xlsx' if content[:4] == _ZIP_MAGIC else 'unknown')
    print(f"[upload] fname={fname!r}, magic={content[:4].hex()}, detected={actual_fmt}")

    # Convert OLE2 (xls) regardless of declared extension
    if actual_fmt == 'xls':
        try:
            content = _xls_to_xlsx(content)
        except Exception as exc:
            exc_msg = str(exc)
            if "workbook" in exc_msg.lower() or "OLE2" in exc_msg:
                raise HTTPException(
                    status_code=422,
                    detail=(
                        "File is encrypted or IRM-protected. "
                        "To fix: open in Excel, select-all (Ctrl+A), copy, paste into a NEW blank workbook, "
                        "save as .xlsx, then upload that new file."
                    ),
                )
            raise HTTPException(status_code=422, detail=f"Cannot convert .xls: {exc}")
        if fname.lower().endswith(".xls"):
            fname = fname[:-4] + ".xlsx"
    elif actual_fmt == 'unknown':
        raise HTTPException(status_code=422, detail="File does not appear to be a valid Excel file (not OLE2 or zip/xlsx format)")

    # Save file to disk
    reg_dir = DATA_DIR / "registers"
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S_%f")
    dest = reg_dir / f"{ts}_{fname}"
    dest.write_bytes(content)

    # Parse to get counts (use in-memory bytes to avoid antivirus file lock)
    try:
        import io as _io
        registers, bitfields = parse_excel(_io.BytesIO(content))
    except Exception as exc:
        dest.unlink(missing_ok=True)
        raise HTTPException(status_code=422, detail=f"Cannot parse Excel: {exc}")

    effective_name = name.strip() if name and name.strip() else Path(fname).stem

    record = RegisterDefinitionORM(
        name=effective_name,
        original_filename=fname,
        file_path=str(dest),
        register_count=len(registers),
        bitfield_count=len(bitfields),
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@router.patch("/{register_id}", response_model=RegisterDefinitionOut)
def rename_register(register_id: int, payload: RegisterRenameIn, db: Session = Depends(get_db)):
    record = db.get(RegisterDefinitionORM, register_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Register not found")
    new_name = (payload.name or "").strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="Name cannot be empty")
    record.name = new_name
    db.commit()
    db.refresh(record)
    return record


@router.delete("/{register_id}", status_code=204)
def delete_register(register_id: int, db: Session = Depends(get_db)):
    record = db.get(RegisterDefinitionORM, register_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Register not found")

    if record.batches:
        raise HTTPException(
            status_code=409,
            detail="Cannot delete: there are batches referencing this register definition",
        )

    file_path = Path(record.file_path)
    db.delete(record)
    db.commit()
    file_path.unlink(missing_ok=True)
